import pyodbc
import openpyxl

from helper import dbg
from openpyxl import Workbook
from openpyxl.styles import Font
from pyodbc import Connection, Cursor
from openpyxl.worksheet.worksheet import Worksheet


class Column:
    def __init__(self, constraint, column_name, data_type, max_length, required):
        self.constraint_type = ""
        self.constraint_value = ""

        if constraint:
            constraint_list: list[str] = constraint.split('_')
            self.constraint_type: str = constraint_list[0]

            if self.constraint_type == 'PK':
                self.constraint_value = f"Уникальный идентификатор таблицы «{constraint_list[1]}»"
            elif self.constraint_type == 'FK':
                self.constraint_value = f"Ссылается на таблицу «{constraint_list[2]}»"
        self.column_name: str = column_name

        self.max_length: str = 'MAX' if max_length == -1 else str(max_length)
        self.data_type: str = data_type.upper() if max_length is None else f"{data_type.upper()}({self.max_length})"

        self.required: str = {"YES": "N", "NO": "Y"}[required]

    def __str__(self):
        return f"{self.constraint_type:2} | {self.column_name:20} | {self.data_type:20} | {self.required:1} | {self.constraint_value}"


class DataDictionary:
    def __init__(self, server, db_name):
        self.__db_name: str = db_name
        self.__connection: Connection = pyodbc.connect(r"Driver={SQL Server};"
                                                       fr"Server={server};"
                                                       fr"Database={db_name};"
                                                       r"Trusted_Connection=yes;")

        self.__cursor: Cursor = self.__connection.cursor()
        dbg("connection is open")
        self.__wb: Workbook = openpyxl.Workbook()
        self.__ws: Worksheet | None = self.__wb.active
        dbg("excel is open")

    def get_table(self):
        # Запросы
        query_tables = "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME != 'sysdiagrams'"
        query_columns = (
            "SELECT CONSTRAINT_NAME, INFORMATION_SCHEMA.COLUMNS.COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, IS_NULLABLE\n" 
            "FROM INFORMATION_SCHEMA.COLUMNS\n"
            "LEFT JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE\n"
            "ON INFORMATION_SCHEMA.COLUMNS.COLUMN_NAME = INFORMATION_SCHEMA.KEY_COLUMN_USAGE.COLUMN_NAME\n"
            "WHERE INFORMATION_SCHEMA.COLUMNS.TABLE_NAME = '?' and (CONSTRAINT_NAME LIKE 'FK[_]?[_]%' or CONSTRAINT_NAME LIKE 'PK[_]?' or CONSTRAINT_NAME is NULL)")

        # Получаем таблицы
        tables: list[str] = list(i for i in map(lambda _: _[0], self.__cursor.execute(query_tables).fetchall()) if not i.startswith('_'))

        # Дефолтные значения
        row: int = 1
        pattern: list[str] = ["KEY", "FIELD NAME", "DATA TYPE/FIELD SIZE", "REQUIRED?", "NOTES"]

        # Проходимся по всем таблицам
        for table in tables:
            # Добавляем название таблицы
            self.__ws.cell(row=row, column=1).value = table
            self.__ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            # Добавляем названия столбцов для таблицы
            for index, row_title in enumerate(pattern):
                self.__ws.cell(row=row, column=index + 1).value = row_title
                self.__ws.cell(row=row, column=index + 1).font = Font(bold=True)
            row += 1

            # Получаем строки из запроса
            columns: list[Column] = list(map(lambda _: Column(*_),
                                             self.__cursor.execute(query_columns.replace('?', table)).fetchall()))

            # Выводим
            print(table, end='\n' + '-' * 120 + '\n')
            print(*columns, sep='\n', end='\n' + '-' * 120 + '\n' * 2)

            # Наполняем столбцы значениями
            for column in columns:
                # KEY
                self.__ws.cell(row=row, column=1).value = column.constraint_type

                # FIELD NAME
                self.__ws.cell(row=row, column=2).value = column.column_name

                # DATA TYPE/FIELD SIZE
                self.__ws.cell(row=row, column=3).value = column.data_type

                # REQUIRED?
                self.__ws.cell(row=row, column=4).value = column.required

                # NOTES
                self.__ws.cell(row=row, column=5).value = column.constraint_value

                # Переход
                row += 1
            row += 1

        # Сохранение изменений
        self.__wb.save(f"{self.__db_name.lower()}_data_dict.xlsx")
        dbg("excel saved")

    def __del__(self):
        self.__wb.close()
        dbg("excel is closed")
        self.__connection.close()
        dbg("connection is closed")
